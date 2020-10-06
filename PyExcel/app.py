import openpyxl

# wb = openpyxl.Workbook()
wb = openpyxl.load_workbook("PyExcel/transactions.xlsx")
print(wb.sheetnames)

sheet = wb["Sheet1"]

cell = sheet["a1"]
column = sheet["a"]
cells = sheet["a:c"]
sheet["a1:c3"]
sheet[1:3]

sheet.append([1, 2, 3])
wb.save("PyExcel/transactions2.xlsx")